from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes,authentication_classes
from rest_framework.permissions import AllowAny,IsAuthenticated
from .models import Cuota, Profile,User,Product, Order, OrderProduct
from .serializers import AdminUserCreationSerializer, CuotaSerializer, UserSerializer,ProfileSerializer,ProductSerializer, OrderSerializer
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import check_password
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework import viewsets,permissions
import json
from django.http import JsonResponse
from rest_framework.permissions import IsAdminUser
from django.shortcuts import get_object_or_404, render
from rest_framework import status
from django.contrib.auth import authenticate
from rest_framework.views import APIView
from openpyxl import Workbook
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
from decimal import Decimal
import mercadopago
import logging
from django.conf import settings
from django.core.mail import send_mail
from rest_framework.permissions import AllowAny
import re
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from rest_framework_simplejwt.authentication import JWTAuthentication
from .serializers import ComprobantePagoSerializer


@csrf_exempt
def contacto(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            nombre = data.get('nombre')
            email = data.get('email')
            asunto = data.get('asunto')
            mensaje = data.get('mensaje')

            if not all([nombre, email, asunto, mensaje]):
                return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios'}, status=400)

            cuerpo_mensaje = f"""
            Has recibido un nuevo mensaje desde el formulario de contacto:

            Nombre: {nombre}
            Correo electrónico: {email}
            Asunto: {asunto}
            Mensaje:
            {mensaje}
            """

            send_mail(
                subject=f"Nuevo mensaje: {asunto}",
                message=cuerpo_mensaje,
                from_email=email,
                recipient_list=['contacto.cooperadora.escuela@gmail.com'],
                fail_silently=False,
            )

            return JsonResponse({'success': True, 'message': 'Mensaje enviado correctamente'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)



# preferencia de pago

#fin preferencia de pago


# vits usurio
class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(is_superuser=False)
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
#fin vits usurio
    
    
# permisos para el admin
class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_staff
#fin permisos para el admin  

 
# perfil de usuario  
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])  #solo usuarios autenticados pueden acceder
def profile_view(request):
    user = request.user 
    try: 
        profile = request.user.profile  
    except Profile.DoesNotExist:
        return Response({'error': 'Perfil no encontrado'}, status=404)

    if request.method == 'GET':
        #muestra los datos del perfil
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)

    elif request.method == 'PUT':
        #actualiza los datos del perfil
        serializer = ProfileSerializer(profile, data=request.data, partial=True)  
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
#fin perfil de usuario  


# iniciar sesion
User = get_user_model()
@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])  
def login(request):
    email = request.data.get('email')
    password = request.data.get('password')

    if not email or not password:
        return Response({'error': 'Email y contraseña son requeridos'}, status=400)

    try:
        #verifica si esta en la base de datos
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response({'error': 'Usuario no encontrado'}, status=404)

    #verifica si la contraseña es correcta
    if not check_password(password, user.password):
        return Response({'error': 'Contraseña incorrecta'}, status=401)

    #verificar si el usuario esta activo
    if not user.is_active:
        return Response({'error': 'Cuenta inactiva'}, status=403)

    #genera los tokens jwt
    refresh = RefreshToken.for_user(user)

    #devuelve los tokens y la info del usuario
    return Response({
        'refresh': str(refresh),
        'access': str(refresh.access_token),
        'user': {
            'id': user.id,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'is_staff': user.is_staff,          
            'is_superuser': user.is_superuser 
        }
    })
#fin iniciar sesion


# validar contraseña
def secure_password(password):
    if len(password) < 8:
        return "La contraseña debe tener al menos 8 caracteres."
    if not re.search(r'[A-Z]', password):
        return "Debe incluir al menos una letra mayúscula."
    if not re.search(r'[@$!%*?&._-]', password):
         return "Debe incluir al menos un carácter especial (@$!%*?&._-)."
    return None
# fin contraseña

# registro
User = get_user_model()
@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            first_name = data.get('first_name')
            last_name = data.get('last_name')
            email = data.get('email')
            password = data.get('password')
            password2 = data.get('password2')

            #verifica que no falte datos
            if not all([first_name, last_name, email, password, password2]):
                return JsonResponse({"error": "Todos los campos son obligatorios."}, status=400)

            #que las contraseñas sean iguales
            if password != password2:
                return JsonResponse({"error": "Las contraseñas no coinciden."}, status=400)
            
            #  contraseña segura
            error_password = secure_password(password)
            if error_password:
                return JsonResponse({"error": error_password}, status=400)


            #si el email ya existe
            if User.objects.filter(email=email).exists():
                return JsonResponse({"error": "El email ya está registrado."}, status=400)

            #se crea el usuario
            user = User.objects.create_user(
                username=email,  #usar el email como username
                first_name=first_name,
                last_name=last_name,
                email=email,
                password=password
            )

            #crea el perfil
            Profile.objects.create(user=user)

            return JsonResponse({"message": "Usuario creado exitosamente."}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Formato JSON inválido."}, status=400)

        except Exception as e:
            return JsonResponse({"error": f"Error en el servidor: {str(e)}"}, status=500)

    else:
        return JsonResponse({"error": "Método no permitido."}, status=405)
#fin registro

from rest_framework.parsers import MultiPartParser, FormParser
# producto
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    def update(self, request, pk=None):
        product = get_object_or_404(Product, pk=pk)
        product.name = request.data.get('name', product.name)
        product.price = request.data.get('price', product.price)
    
        if 'image' in request.FILES:
            product.image = request.FILES['image']  

        product.save()
        return Response({'message': 'Producto actualizado correctamente'})

#fin producto


# orden
class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
# fin orden


# checkout
   
logger = logging.getLogger(__name__)

class CheckoutView(APIView):
    def post(self, request):
        serializer = OrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                order = serializer.save()
                logger.info("Orden guardada")

                sdk = mercadopago.SDK(settings.MERCADOPAGO_ACCESS_TOKEN)

                items = []
                for order_product in OrderProduct.objects.filter(order=order):
                    items.append({
        'title': order_product.product.name,
        'quantity': order_product.quantity,
        'unit_price': float(order_product.unit_price),
        'currency_id': 'ARS'
    })

                logger.info(f"Items: {items}")

                preference_data = {
                    "items": items,
                    "back_urls": {
                        "success": "http://localhost:4200/success",
                        "failure": "http://localhost:4200/failure",
                        "pending": "http://localhost:4200/pending"
                    }
                }

                preference_response = sdk.preference().create(preference_data)
                logger.info(f"Respuesta MercadoPago: {preference_response}") #muestra el error por consola

                if preference_response["status"] == 201:
                    preference_url = preference_response["response"]["init_point"]
                    logger.info(f"URL de pago: {preference_url}")  
                    return Response({"payment_url": preference_url}, status=status.HTTP_201_CREATED)
                else:
                    logger.error(f"Error en la creación de la preferencia: {preference_response}")
                    return Response(preference_response["response"], status=status.HTTP_400_BAD_REQUEST)

            except Exception as e:
                logger.error(f"Error en CheckoutView: {str(e)}")
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
# fin checkout  

# status
@csrf_exempt
def payment_status(request):
    payment_id = request.GET.get("payment_id")  

    if not payment_id:
        return JsonResponse({"error": "No se proporcionó payment_id"}, status=400)

    try:
        sdk = mercadopago.SDK(settings.MERCADOPAGO_ACCESS_TOKEN)

        payment_info = sdk.payment().get(payment_id)

        if payment_info["status"] == 200:
            status_pago = payment_info["response"]["status"]
            preference_id = payment_info["response"]["order"]["id"]  

            
            try:
                order = Order.objects.get(preference_id=preference_id)
                if status_pago == "approved":
                    order.status = "pagado"
                    order.save()
            except Order.DoesNotExist:
                return JsonResponse({"error": "Orden no encontrada"}, status=404)

            return JsonResponse({"payment_status": status_pago})
        else:
            return JsonResponse({"error": "Error al consultar el pago"}, status=payment_info["status"])

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
# fin status     
    

# descargar excel 
def download_orders_excel(request):
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes"

    # Agregar encabezados
    headers = [
        "ID", 
        "Nombre", 
        "Apellido", 
        "DNI", 
        "Producto", 
        "Cantidad", 
        "Precio Unitario", 
        "Forma de Pago"
    ]
    ws.append(headers)

    # Estilo para los encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Obtener todas las órdenes y calcular el total general
    orders = Order.objects.all()
    total_general = Decimal('0.0')  # Usar Decimal para el total general

    for order in orders:
        for order_product in order.orderproduct_set.all():
            precio_unitario = order_product.unit_price  # Precio unitario (Decimal)
            total_general += precio_unitario  # Sumar al total general

            ws.append([
                order.id,
                order.name,
                order.surname,
                order.dni,
                order_product.product.name,
                order_product.quantity,
                float(precio_unitario),  # Convertir a float para el archivo Excel
                order.payment_method
            ])

    # Agregar una fila con el total general
    ws.append([])  # Fila vacía para separar
    total_row = [
        "",  # ID
        "",  # Nombre
        "",  # Apellido
        "",  # DNI
        "",  # Producto
        "",  # Cantidad
        "Total General",  # Etiqueta
        float(total_general)  # Valor del total general (convertido a float)
    ]
    ws.append(total_row)

    # Aplicar estilo a la fila del total general
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")  # Texto en blanco y negrita
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Fondo azul

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Guardar el archivo en memoria
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ordenes.xlsx'

    return response
# fin descargar excel 



# vista para que el admin realice el crud de usuarios
@api_view(['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAdminUser])
def all_profile_view(request, pk=None):
    if request.method == 'GET':
        if pk:
            try:
                profile = Profile.objects.get(pk=pk)
                serializer = ProfileSerializer(profile)
                return Response(serializer.data)
            except Profile.DoesNotExist:
                return Response({'error': 'Perfil no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        else:
            profiles = Profile.objects.all()
            serializer = ProfileSerializer(profiles, many=True)
            return Response(serializer.data)


    elif request.method == 'PUT':
        try:
            profile = Profile.objects.get(pk=pk)
        except Profile.DoesNotExist:
            return Response({'error': 'Perfil no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():

            user = profile.user
            user.first_name = serializer.validated_data.get('first_name', user.first_name)
            user.last_name = serializer.validated_data.get('last_name', user.last_name)
            user.save()

            # Guardamos el perfil
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        try:
            profile = Profile.objects.get(pk=pk)
        except Profile.DoesNotExist:
            return Response({'error': 'Perfil no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        user = profile.user
        profile.delete()
        user.delete()
        return Response({'message': 'Perfil y usuario eliminados exitosamente'}, status=status.HTTP_204_NO_CONTENT)
    
@api_view(['POST'])
@permission_classes([IsAdminUser])
def register_user_by_admin(request):
    serializer = AdminUserCreationSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({"message": "Usuario creado exitosamente."}, status=201)
    return Response(serializer.errors, status=400)
    
# fin vista para que el admin realice el crud de usuarios


# tramite
# class ProcedureViewSet(viewsets.ModelViewSet):
#     serializer_class = ProcedureSerializer
#     permission_classes = [permissions.IsAuthenticated]

#     def get_queryset(self):
#         return Procedure.objects.filter(user=self.request.user)  # Cambié 'usuario' por 'user'

#     def perform_create(self, serializer):
#         serializer.save(user=self.request.user)
# fin tramite

# cuota
@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def descargar_comprobante(request, cuota_id):
    cuota = get_object_or_404(Cuota, id=cuota_id, usuario=request.user)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=comprobante_cuota_{cuota.id}.pdf'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width / 2, height - 50, "Comprobante de Cuota Cooperadora Escuela")
    p.setLineWidth(1)
    p.setStrokeColor(colors.grey)
    p.line(30, height - 60, width - 30, height - 60)

    # Datos del usuario
    p.setFont("Helvetica", 12)
    y = height - 100
    line_height = 20

    p.drawString(50, y, f"Nombre del alumno: {cuota.usuario.get_full_name() or cuota.usuario.username}")
    y -= line_height
    tipo = cuota.get_tipo_display()
    mes = dict(Cuota.MESES_CHOICES).get(cuota.mes, '') if cuota.mes else ''
    p.drawString(50, y, f"Tipo de cuota: {tipo}")
    y -= line_height
    if mes:
        p.drawString(50, y, f"Mes: {mes}")
        y -= line_height
    p.drawString(50, y, f"Año: {cuota.anio}")
    y -= line_height
    p.drawString(50, y, f"Monto: ${cuota.monto}")
    y -= line_height
    p.drawString(50, y, f"Estado de pago: {'PAGADO' if cuota.pagado else 'PENDIENTE'}")
    y -= line_height
    p.drawString(50, y, f"Fecha de generación: {cuota.fecha_creacion.strftime('%d/%m/%Y')}")
    y -= line_height
    p.drawString(50, y, f"Número de comprobante: {cuota.nro_comprobante}")
    y -= line_height

    
    y -= 20
    p.line(30, y, width - 30, y)
    y -= 30

    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Datos para transferencia bancaria:")
    y -= line_height
    p.setFont("Helvetica", 12)
    p.drawString(70, y, "CBU: 0000003100001234567890")
    y -= line_height
    p.drawString(70, y, "Alias: cooperadora.escuela")
    y -= line_height
    p.drawString(70, y, "Banco: Banco Nación")
    
    y -= 30
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(colors.black)
    p.drawString(50, y, "IMPORTANTE:")
    y -= line_height
    p.setFont("Helvetica", 11)
    p.drawString(50, y, "Una vez abonado, enviar el número de transferencia al email:")
    y -= line_height
    p.drawString(50, y, "cooperadora@escuela.edu.ar o al formulario en la sección pagar cuota")


    p.setFont("Helvetica-Oblique", 10)
    p.setFillColor(colors.grey)
    p.drawCentredString(width / 2, 120, "Gracias por colaborar con la cooperadora escolar.")

    p.showPage()
    p.save()
    return response


# calculo el monto segun cuota
def calcular_monto(tipo):
    return {
        'mensual': 1000,
        'anual': 10000
    }.get(tipo, 0)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_cuota(request):
    data = request.data
    tipo = data.get('tipo')
    mes = data.get('mes') 
    anio = data.get('anio')

    if not tipo or tipo not in dict(Cuota.TIPO_CHOICES):
        return Response({'error': 'Tipo inválido'}, status=status.HTTP_400_BAD_REQUEST)
    
    if not anio:
        return Response({'error': 'Año requerido'}, status=status.HTTP_400_BAD_REQUEST)

    if tipo == 'mensual' and not mes:
        return Response({'error': 'Mes requerido para cuota mensual'}, status=status.HTTP_400_BAD_REQUEST)

    monto = calcular_monto(tipo)

    cuota = Cuota.objects.create(
        usuario=request.user,
        tipo=tipo,
        mes=mes if tipo == 'mensual' else None,
        anio=anio,
        monto=monto
    )

    serializer = CuotaSerializer(cuota)
    return Response(serializer.data, status=status.HTTP_201_CREATED)

import qrcode
from io import BytesIO


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def ver_qr_pago(request, cuota_id):
    cuota = get_object_or_404(Cuota, id=cuota_id, usuario=request.user)

    
    datos_qr = f"""
    Nombre: {cuota.usuario.get_full_name() or cuota.usuario.username}
    Cuota: {cuota.get_tipo_display()} {cuota.mes or ''}/{cuota.anio}
    Monto: ${cuota.monto}
    CBU: 0000003100001234567890
    Alias: cooperadora.escuela
    """

    # genero el qr
    qr = qrcode.make(datos_qr)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    buffer.seek(0)

    return HttpResponse(buffer.getvalue(), content_type="image/png")






class EnviarComprobanteView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        serializer = ComprobantePagoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(user=request.user)  # <-- asignas el usuario aquí
            return Response({'mensaje': 'Comprobante guardado correctamente'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
